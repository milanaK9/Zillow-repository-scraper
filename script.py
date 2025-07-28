import random
import time
from marshal import dumps

import requests
import playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

import pyzill
import json


from playwright.sync_api import sync_playwright
import json


import urllib.parse
import json
from typing import Any, Optional
from curl_cffi import requests
from playwright.sync_api import sync_playwright

sq = "los angeles ca"

def get_zillow_json(url):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=50)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            locale="en-US",
            viewport={"width": 1280, "height": 800}
        )

        # Anti-bot stealth tweaks
        context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        page = context.new_page()
        print(f"🔍 Visiting: {url}")
        page.goto(url)
        # page.wait_for_load_state("networkidle")

        # Allow page scripts to finish
        time.sleep(3)

        try:
            raw_json = page.query_selector('#__NEXT_DATA__').text_content()
            zillow_data = json.loads(raw_json)

            return zillow_data

        except Exception as e:
            print(f"❌ Error extracting searchQueryState: {e}")
            return None

        finally:
            browser.close()
def get_info(search_query):
    zillow_data = get_zillow_json(f"https://www.zillow.com/{search_query.strip().replace(' ', '-').lower()}")
    #search_query_state = zillow_data["props"]["pageProps"]["searchPageState"]['queryState']
    total_pages = zillow_data["props"]["pageProps"]["searchPageState"]['cat1']['searchList']['totalPages']
    print("✅ totalPages extracted!")
    return  total_pages


props = []
total_pages = get_info(sq)

for i in range(total_pages):
    time.sleep(random.uniform(2, 5))
    url = "https://www.zillow.com/" + sq.strip().replace(' ', '-').lower() + "/" \
           + str(i + 1) + "_p" + "/"

    zillow_data = get_zillow_json(url)
    for prop in zillow_data["props"]["pageProps"]["searchPageState"]['cat1']['searchResults']['listResults']:
        props.append({
            "Price": prop.get("price"),
            "Address": prop.get("address"),
            "Bedrooms": prop.get("beds", "N/A"),
            "Bathrooms": prop.get("baths", "N/A"),
            "Living Area (sqft)": str(prop.get("area")) + " sqft" if prop.get("area", None) else "N/A",
            "Latitude": prop.get("latLong", {}).get("latitude"),
            "Longitude": prop.get("latLong", {}).get("longitude"),
            "Property Link": prop.get("detailUrl", ""),
        })

for prop in props:
    price = prop.get("Price")
    # Sometimes price can be a string with $ or a number, handle both
    if not price or str(price).strip() in ["$0", "0", "0.0", 0]:
        prop["Price"] = "N/A"

import pandas as pd

df = pd.DataFrame(props)
df.to_excel("properties.xlsx", index=False)
def format_excel_file(filename="properties.xlsx"):
    wb = load_workbook(filename)
    ws = wb.active

    # Bold and center headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Auto-fit columns and make Property Link clickable
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        header = col_cells[0].value.lower() if col_cells[0].value else ""

        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            if header == "property link" and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"
            max_len = max(max_len, len(val))

        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filename)
    print(f"✅ Excel file formatted and saved: {filename}")

format_excel_file("properties.xlsx")
import csv

keys = props[0].keys()  # column headers

with open("properties.csv", "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=keys)
    writer.writeheader()
    writer.writerows(props)
